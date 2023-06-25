import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
import bank_system

def validate_login(username, password):
    try:
        workbook = load_workbook('database/customers.xlsx')
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[2]) == username and str(row[3]) == password:
                return True

        return False
    except FileNotFoundError:
        messagebox.showerror("Error", "Data file not found.")
        return False

def login():
    username = username_entry.get()
    password = password_entry.get()

    if validate_login(username, password):
        messagebox.showinfo("Success", "Login successful!")
        window.destroy()
        bank_system.open_bank_system(username)
    else:
        messagebox.showerror("Error", "Wrong details.")

window = tk.Tk()
window.title("Bank App")

username_label = tk.Label(window, text="Username:")
username_label.pack()
username_entry = tk.Entry(window)
username_entry.pack()

password_label = tk.Label(window, text="Password:")
password_label.pack()
password_entry = tk.Entry(window, show="*")
password_entry.pack()

login_button = tk.Button(window, text="Login", command=login)
login_button.pack()

window.mainloop()
