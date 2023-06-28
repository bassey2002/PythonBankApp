import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
import customers

# Global variables
current_user = None
window = None
amount_entry = None

def deposit():
    amount = float(amount_entry.get())
    # Update the "Amount" column in customers.xlsx with the deposited amount
    update_balance(amount)
    messagebox.showinfo("Deposit Successful", "Deposit successful!")
    print_balance()

def withdraw():
    amount = float(amount_entry.get())
    # Check if the withdrawal amount is valid
    if validate_withdrawal(amount):
        # Subtract the amount from the "Amount" column in customers.xlsx
        update_balance(-amount)
        messagebox.showinfo("Withdrawal Successful", "Withdrawal successful!")
        print_balance()
    else:
        messagebox.showerror("Invalid Withdrawal", "Insufficient balance!")

def check_balance():
    # Retrieve the current balance from the "Amount" column in customers.xlsx
    balance = get_balance()
    messagebox.showinfo("Current Balance", f"Your current balance is: {balance}")

def logout():
    global current_user
    current_user = None
    window.destroy()
    # Reopen the login page (customers.py)
    customers.window.deiconify()

def update_balance(amount):
    try:
        workbook = load_workbook('database/customers.xlsx')
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[2]) == current_user:
                new_balance = row[5] + amount
                row_list = list(row)  # Convert tuple to list
                row_list[5] = new_balance  # Modify the element
                updated_row = tuple(row_list)  # Convert back to tuple
                sheet.append(updated_row)  # Append the updated row

        workbook.save('database/customers.xlsx')
        workbook.close()
    except FileNotFoundError:
        messagebox.showerror("Error", "Data file not found.")


def validate_withdrawal(amount):
    try:
        workbook = load_workbook('database/customers.xlsx')
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[2]) == current_user:
                balance = row[5]
                if balance >= amount:
                    workbook.close()
                    return True

        workbook.close()
        return False
    except FileNotFoundError:
        messagebox.showerror("Error", "Data file not found.")
        return False

def get_balance():
    try:
        workbook = load_workbook('database/customers.xlsx')
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[2]) == current_user:
                balance = row[5]
                workbook.close()
                return balance

        workbook.close()
        return 0
    except FileNotFoundError:
        messagebox.showerror("Error", "Data file not found.")
        return 0

def print_balance():
    balance = get_balance()
    print(f"Your current balance is: {balance}")

def open_bank_system(user):
    global current_user, window, amount_entry
    current_user = user

    window = tk.Tk()
    window.title("Bank System")

    amount_label = tk.Label(window, text="Enter Amount:")
    amount_label.pack()
    amount_entry = tk.Entry(window)
    amount_entry.pack()

    deposit_button = tk.Button(window, text="Deposit", command=deposit)
    deposit_button.pack()

    withdraw_button = tk.Button(window, text="Withdraw", command=withdraw)
    withdraw_button.pack()

    balance_button = tk.Button(window, text="Check Balance", command=check_balance)
    balance_button.pack()

    logout_button = tk.Button(window, text="Logout", command=logout)
    logout_button.pack()

    window.mainloop()

# Entry point for the application
if __name__ == "__main__":
    customers.login_window.title("Bank App")
    login_button = tk.Button(customers.login_window, text="Login", command=customers.login)
    login_button.pack()

    customers.login_window.mainloop()
